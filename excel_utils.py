# excel_utils.py - Processa arquivos XLSX para extração das informações  
# Autor: Hercules Monteiro
# Data: 26/03/2025
# Versão: 1.0
# ============================================================================
# 1. IMPORTAÇÕES NECESSÁRIAS
# ============================================================================
from openpyxl import load_workbook
import logging
import os

# Configuração do logger para este módulo
logger = logging.getLogger(__name__)

def process_excel(file_path: str) -> list:
    """
    Processa um arquivo Excel e extrai o texto de todas as células.
    
    Args:
        file_path (str): Caminho completo para o arquivo Excel.
    
    Returns:
        list: Uma lista onde cada elemento é o texto extraído de uma linha do Excel.
    """
    try:
        # Verifica se o arquivo existe
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"O arquivo Excel não foi encontrado: {file_path}")
        
        # Carrega o workbook do Excel
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Lista para armazenar o texto de cada linha
        text = []
        
        # Itera sobre todas as planilhas do workbook
        for sheet in wb:
            # Itera sobre todas as linhas da planilha
            for row in sheet.iter_rows():
                # Extrai o texto de cada célula na linha, ignorando células vazias
                cells_texts = [str(cell.value) for cell in row if cell.value is not None]
                # Junta o texto das células em uma única string e adiciona à lista
                text.append(' '.join(cells_texts))
        
        # Retorna a lista com o texto de todas as linhas
        return text
    
    except Exception as e:
        # Captura e registra qualquer erro que possa ocorrer durante o processamento
        logger.error(f"Erro ao processar o arquivo Excel {file_path}: {str(e)}", exc_info=True)
        # Retorna uma lista vazia em caso de erro
        return []

def get_excel_metadata(file_path: str) -> dict:
    """
    Extrai metadados básicos de um arquivo Excel.
    
    Args:
        file_path (str): Caminho completo para o arquivo Excel.
    
    Returns:
        dict: Um dicionário contendo os metadados do arquivo Excel.
    """
    try:
        # Carrega o workbook do Excel
        wb = load_workbook(file_path, read_only=True)
        
        # Extrai informações básicas
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)
        
        # Cria um dicionário com os metadados
        metadata = {
            "filename": os.path.basename(file_path),
            "total_sheets": total_sheets,
            "sheet_names": sheet_names,
            "file_size": os.path.getsize(file_path)
        }
        
        return metadata
    except Exception as e:
        logger.error(f"Erro ao extrair metadados do Excel {file_path}: {str(e)}", exc_info=True)
        return {}

def is_valid_excel(file_path: str) -> bool:
    """
    Verifica se um arquivo é um Excel válido.
    
    Args:
        file_path (str): Caminho completo para o arquivo a ser verificado.
    
    Returns:
        bool: True se o arquivo for um Excel válido, False caso contrário.
    """
    try:
        load_workbook(file_path, read_only=True)
        return True
    except Exception as e:
        logger.error(f"Erro ao validar o arquivo Excel {file_path}: {str(e)}", exc_info=True)
        return False
